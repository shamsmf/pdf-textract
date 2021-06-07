import boto3
import time
import sys
from openpyxl import load_workbook, Workbook
import re
import os

# Create Amazon S3 resource
s3 = boto3.resource('s3')

s3BucketName = 'xxxxxxxxxxxxxxxxxxxxxxxxxxxxx'
documentName = ''
filePath = input('\n Enter a pdf file path:')

if not filePath or filePath is None:
    print('File is missing.')
    sys.exit()
documentName = filePath.split('/')[-1]

# Uploading to s3 bucket
try:
    data = open(filePath, 'rb')
    s3.Bucket(s3BucketName).put_object(Key=documentName, Body=data)
except Exception as e:
    print('Unable to upload file into S3 bucket.')
    sys.exit()


def startJob(s3BucketName, objectName):
    '''Start the jon to extract the content'''
    response = None

    # Start the text detection, of the object from the S3 bucket.
    try:
        client = boto3.client('textract')
        response = client.start_document_text_detection(
            DocumentLocation={
            'S3Object': {
            'Bucket': s3BucketName,
            'Name': objectName
            }
            })
    except Exception as e:
        print('Unable to start text detection')
        sys.exit()
    
    return response["JobId"]

def isJobComplete(jobId):
    '''Check the job status'''
    time.sleep(5)

    # Use jobid to check the status, it will continue checking until the status is SUCCEEDED
    try:
        client = boto3.client('textract')
        response = client.get_document_text_detection(JobId=jobId)
        status = response["JobStatus"]
        print("Job status: {}".format(status))

        while(status == "IN_PROGRESS"):
            time.sleep(5)
            response = client.get_document_text_detection(JobId=jobId)
            status = response["JobStatus"]
            print("Job status: {}".format(status))
    except Exception as e:
        print('Jon failed');
        sys.exit()

    return status

def getJobResults(jobId):
    '''return the result from aws textract'''

    pages = []

    time.sleep(5)

    # Get the pdf detected reposnse from aws textract service
    try:
        client = boto3.client('textract')
        response = client.get_document_text_detection(JobId=jobId)
        pages.append(response)
        print("Resultset page recieved: {}".format(len(pages)))
        nextToken = None

        if('NextToken' in response):
            nextToken = response['NextToken']

            while(nextToken):
                time.sleep(5)
                response = client.get_document_text_detection(JobId=jobId, NextToken=nextToken)
                pages.append(response)
                print("Resultset page recieved: {}".format(len(pages)))
                nextToken = None

                if('NextToken' in response):
                    nextToken = response['NextToken']
    except Exception as e:
        print('Failed job to fetch response.')
        sys.exit()
    
    return pages

# getting jobid
jobId = startJob(s3BucketName, documentName)

print('Job status: STARTED')
if(isJobComplete(jobId)):
    response = getJobResults(jobId)

# Predefined fields in the pdf form
keys = {
'ADDRESS OF PREMISE': ['ADDRESS OF PREMISE', 'THIS AGREEMENT'],
'IN THE PRESENCEOF': ['IN THE PRESENCEOF', ' ']
}

output = []
keyFound = ''

# Create a text from the lines in the response.
formText = ''
for resultPage in response:
    for item in resultPage['Blocks']:
        if item["BlockType"] == 'LINE':
            formText += str(item['Text']) + ' '
               
# find the searchable key values
for key in keys:
    #start index of the search key
    startVal = re.search(keys[key][0], formText).end() 
    newFormText = formText[startVal+1::]
    # end index of the search key
    endVal = re.search(keys[key][1], newFormText).start() 
    output.append(newFormText[:endVal])

# Create xlsx file if doesn't exists already or load and store the data into xlsx file
outputSheet = './input_sheet_output.xlsx'
if not os.path.isfile(outputSheet):
    headers       = ['ADDRESS OF PREMISE', 'IN THE PRESENCEOF']
    wb = Workbook()
    page = wb.active
    page.title = 'Result'
    page.append(headers)
    wb.save(outputSheet)

wb = load_workbook(outputSheet)
ws = wb.worksheets[0]
ws.append(output)
wb.save(outputSheet)

print('File Path: ' + outputSheet)
